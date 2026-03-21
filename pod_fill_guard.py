import argparse
import json
import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "日期",
    "抽查路线",
    "抽查司机",
    "运单数量",
    "未收数量",
    "错扫数量",
    "晚间更新状态",
    "单号",
]

ROUTE_RE = re.compile(r"^IAH\d{2}-[A-Za-z0-9]+(?:-[A-Za-z0-9]+)?$")
STATION_TEAM_RE = re.compile(r"^IAH-[A-Za-z]+$")


@dataclass
class ValidationIssue:
    level: str
    image_name: str
    route: Optional[str]
    message: str


@dataclass
class Record:
    image_name: str
    date: Optional[str]
    route: Optional[str]
    driver: Optional[str]
    total: Optional[int]
    unscanned: Optional[int]
    exceptions: Optional[int]
    station_team: Optional[str]
    task_count: Optional[int]
    review_required: bool
    review_reason: Optional[str]
    source_fields: Dict[str, Any]

    @classmethod
    def from_dict(cls, image_name: str, payload: Dict[str, Any]) -> "Record":
        return cls(
            image_name=image_name,
            date=payload.get("date"),
            route=payload.get("route"),
            driver=payload.get("driver"),
            total=payload.get("total"),
            unscanned=payload.get("unscanned"),
            exceptions=payload.get("exceptions"),
            station_team=payload.get("station_team"),
            task_count=payload.get("task_count"),
            review_required=bool(payload.get("review_required", False)),
            review_reason=payload.get("review_reason"),
            source_fields=payload.get("source_fields") or {},
        )

    def as_excel_row(self) -> List[Any]:
        return [
            self.date,
            self.route,
            self.driver,
            self.total,
            self.unscanned,
            self.exceptions,
            "",
            "",
        ]


def load_records(path: str) -> List[Record]:
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    images: Sequence[Dict[str, Any]]
    if isinstance(payload, dict) and "images" in payload:
        images = payload["images"]
    elif isinstance(payload, list):
        images = payload
    else:
        raise ValueError("Input JSON must be a list or an object with an 'images' field.")

    records: List[Record] = []
    for image in images:
        image_name = image.get("image_name")
        if not image_name:
            raise ValueError("Each image item must include image_name.")
        for record_payload in image.get("records", []):
            records.append(Record.from_dict(image_name, record_payload))
    return records


def init_manifest(images_dir: str, output_path: str) -> None:
    names = sorted(
        name
        for name in os.listdir(images_dir)
        if name.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
    )
    payload = {
        "images": [
            {
                "image_name": name,
                "records": [
                    {
                        "date": None,
                        "route": None,
                        "driver": None,
                        "total": None,
                        "unscanned": None,
                        "exceptions": None,
                        "station_team": None,
                        "task_count": None,
                        "review_required": True,
                        "review_reason": "Pending AI extraction and human review.",
                        "source_fields": {},
                    }
                ],
            }
            for name in names
        ]
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _is_non_negative_int(value: Any) -> bool:
    return isinstance(value, int) and value >= 0


def validate_records(records: Sequence[Record], strict: bool = True) -> List[ValidationIssue]:
    issues: List[ValidationIssue] = []
    seen: set[Tuple[str, str]] = set()

    for record in records:
        key = (record.image_name, record.route or "")
        if key in seen:
            issues.append(
                ValidationIssue("error", record.image_name, record.route, "Duplicate image_name + route pair.")
            )
        seen.add(key)

        if not record.date:
            issues.append(ValidationIssue("error", record.image_name, record.route, "Missing date."))
        if not record.route:
            issues.append(ValidationIssue("error", record.image_name, record.route, "Missing route."))
        elif not ROUTE_RE.match(record.route):
            issues.append(
                ValidationIssue(
                    "error",
                    record.image_name,
                    record.route,
                    "Route format is invalid or may be using station team instead of task route.",
                )
            )

        if not record.driver:
            issues.append(ValidationIssue("error", record.image_name, record.route, "Missing driver name."))

        if not _is_non_negative_int(record.total):
            issues.append(ValidationIssue("error", record.image_name, record.route, "Total must be a non-negative integer."))
        if not _is_non_negative_int(record.unscanned):
            issues.append(
                ValidationIssue("error", record.image_name, record.route, "Unscanned must be a non-negative integer.")
            )
        if not _is_non_negative_int(record.exceptions):
            issues.append(
                ValidationIssue("error", record.image_name, record.route, "Exceptions must be a non-negative integer.")
            )

        if record.station_team and record.route and record.station_team == record.route:
            issues.append(
                ValidationIssue("error", record.image_name, record.route, "Route equals station team; mapping is likely wrong.")
            )

        if record.station_team and not STATION_TEAM_RE.match(record.station_team):
            issues.append(
                ValidationIssue(
                    "warning",
                    record.image_name,
                    record.route,
                    "Station team format looks unusual; double-check top-right field.",
                )
            )

        if record.task_count and record.task_count > 1 and not record.review_required:
            issues.append(
                ValidationIssue(
                    "error",
                    record.image_name,
                    record.route,
                    "Multi-task image must require review unless every task is fully visible and fully extracted.",
                )
            )

        if record.review_required:
            level = "error" if strict else "warning"
            issues.append(
                ValidationIssue(
                    level,
                    record.image_name,
                    record.route,
                    f"Record requires review: {record.review_reason or 'No reason provided.'}",
                )
            )

        driver_text = str(record.source_fields.get("driver_text", "")).strip()
        route_text = str(record.source_fields.get("route_text", "")).strip()
        total_text = str(record.source_fields.get("total_text", "")).strip()
        unscanned_text = str(record.source_fields.get("unscanned_text", "")).strip()
        exceptions_text = str(record.source_fields.get("exceptions_text", "")).strip()

        if driver_text and record.driver and driver_text != record.driver:
            issues.append(
                ValidationIssue("warning", record.image_name, record.route, "Driver differs from source_fields.driver_text.")
            )
        if route_text and record.route and route_text != record.route:
            issues.append(
                ValidationIssue("warning", record.image_name, record.route, "Route differs from source_fields.route_text.")
            )
        if total_text and record.total is not None and total_text != str(record.total):
            issues.append(
                ValidationIssue("warning", record.image_name, record.route, "Total differs from source_fields.total_text.")
            )
        if unscanned_text and record.unscanned is not None and unscanned_text != str(record.unscanned):
            issues.append(
                ValidationIssue(
                    "warning", record.image_name, record.route, "Unscanned differs from source_fields.unscanned_text."
                )
            )
        if exceptions_text and record.exceptions is not None and exceptions_text != str(record.exceptions):
            issues.append(
                ValidationIssue(
                    "warning", record.image_name, record.route, "Exceptions differs from source_fields.exceptions_text."
                )
            )

    return issues


def write_workbook(records: Sequence[Record], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "POD抽查数据"
    ws.append(HEADERS)
    for record in records:
        ws.append(record.as_excel_row())

    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        1: 14,
        2: 18,
        3: 30,
        4: 12,
        5: 12,
        6: 12,
        7: 16,
        8: 20,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output_path)


def write_report(issues: Sequence[ValidationIssue], output_path: str) -> None:
    report = [
        {
            "level": issue.level,
            "image_name": issue.image_name,
            "route": issue.route,
            "message": issue.message,
        }
        for issue in issues
    ]
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


def safe_console_text(value: str) -> str:
    return value.encode("unicode_escape").decode("ascii")


def cmd_init(args: argparse.Namespace) -> int:
    init_manifest(args.images_dir, args.output)
    print(f"Manifest created: {safe_console_text(args.output)}")
    return 0


def cmd_export(args: argparse.Namespace) -> int:
    records = load_records(args.input)
    issues = validate_records(records, strict=not args.allow_review)
    write_report(issues, args.report)

    blocking = [issue for issue in issues if issue.level == "error"]
    if blocking:
        print(f"Blocked: {len(blocking)} error(s). Review report: {safe_console_text(args.report)}")
        return 2

    write_workbook(records, args.output)
    print(f"Workbook created: {safe_console_text(args.output)}")
    if issues:
        print(f"Warnings: {len(issues)}. Review report: {safe_console_text(args.report)}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Guardrail script for POD screenshot extraction, validation, and Excel export."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init", help="Create a manifest JSON from an image folder.")
    init_parser.add_argument("--images-dir", required=True, help="Folder containing POD screenshots.")
    init_parser.add_argument("--output", required=True, help="Output manifest JSON path.")
    init_parser.set_defaults(func=cmd_init)

    export_parser = subparsers.add_parser("export", help="Validate extracted JSON and export Excel.")
    export_parser.add_argument("--input", required=True, help="Input JSON from AI extraction.")
    export_parser.add_argument("--output", required=True, help="Output Excel file path.")
    export_parser.add_argument(
        "--report",
        default="validation_report.json",
        help="Validation report JSON path.",
    )
    export_parser.add_argument(
        "--allow-review",
        action="store_true",
        help="Do not block export on review_required records; use only for supervised workflows.",
    )
    export_parser.set_defaults(func=cmd_export)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
